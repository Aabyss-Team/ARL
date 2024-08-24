#coding: utf-8

from flask import  make_response
from flask_restx import Resource, Namespace
from openpyxl import Workbook
from bson import ObjectId
import re
from collections import Counter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Color
from app.utils import get_logger, auth
from app import utils
from urllib.parse import quote

ns = Namespace('export', description="任务报告导出接口")

logger = get_logger()


@ns.route('/<string:task_id>')
class ARLExport(Resource):
    @auth
    def get(self, task_id):
        """
        报告导出
        """
        task_data = get_task_data(task_id)
        if not task_data:
            return "not found"

        domain = task_data["target"].replace("/", "_")[:20]
        filename = "ARL资产导出报告_{}.xlsx".format(domain)

        excel_data = export_arl(task_id)
        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/octet-stream'
        response.headers["Content-Disposition"] = "attachment; filename={}".format(quote(filename))

        return response




def get_task_data(task_id):
    try:
        task_data = utils.conn_db('task').find_one({'_id': ObjectId(task_id)})
        return task_data
    except Exception as e:
        pass


def get_ip_data(task_id):
    data =  utils.conn_db('ip').find({'task_id': task_id})
    return data


def get_site_data(task_id):
    data = utils.conn_db('site').find({'task_id': task_id})
    return data


def get_domain_data(task_id):
    data = utils.conn_db('domain').find({'task_id': task_id})
    return data


def port_service_product_statist(task_id):
    ip_data = get_ip_data(task_id)
    total = 0
    port_info_list = []
    for item in ip_data:
        if not item["port_info"]:
            continue
        port_info_list.extend(item["port_info"])
        total += len(item["port_info"])

    counter = Counter([info["port_id"] for info in port_info_list])
    top_20 = counter.most_common(20)
    port_percent_list = []
    for port_info in top_20:
        port_id, amount = port_info
        item = {
            "port_id" : port_id,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / total)
        }
        port_percent_list.append(item)

    service_name_list = []
    for info in port_info_list:
        if  not  info.get("product"):
            continue
        if info["product"] or info["version"]:
            service_name = info["service_name"]
            if service_name == "https-alt":
                service_name = "https"

            service_name_list.append(service_name)

    service_top_20 = Counter(service_name_list).most_common(20)
    service_percent_list = []
    for port_info in service_top_20:
        service_name, amount = port_info
        item = {
            "service_name" : service_name,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / len(service_name_list))
        }
        service_percent_list.append(item)



    product_name_list = []
    for info in port_info_list:
        if not info.get("product"):
            continue
        product = info["product"]
        if product and "**" not in product:
            product = product.strip()
            product_name_list.append(product)

    product_top_20 = Counter(product_name_list).most_common(20)
    product_percent_list = []
    for info in product_top_20:
        product, amount = info
        item = {
            "product" : product,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / len(product_name_list))
        }
        product_percent_list.append(item)

    statist = {
        "port_total": total, #端口开放总数
        "port_percent_list": port_percent_list, #端口开放 top 20比例详情
        "service_total": len(service_name_list),  #系统服务类别总数
        "service_percent_list": service_percent_list, #系统服务类别 top 20比例详情
        "product_total": len(product_name_list), #产品种类总数
        "product_percent_list": product_percent_list ##产品种类总数 top 20比例详情
    }
    return statist



class SaveTask(object):
    """docstring for ClassName"""

    def __init__(self, task_id):
        self.task_id = task_id
        self.wb = Workbook()
        self.is_ip_task = False

    def set_style(self, ws):
        font = Font(name="Consolas", color="111111")
        column = "ABCDEFGHIJKLMNO"
        for x in column:
            for y in range(1, 256):
                ws["{}{}".format(x,y)].font = font

    def build_service_xl(self):
        ws = self.wb.create_sheet(title="系统服务")
        ws.column_dimensions['A'].width = 22.0
        ws.column_dimensions['B'].width = 10.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 40.0

        column_tilte = ["IP", "端口","服务", "产品", "版本"]
        ws.append(column_tilte)
        for item in get_ip_data(self.task_id):
            for port_info in item["port_info"]:
                row = []
                row.append(item["ip"])
                row.append("{}".format(port_info["port_id"]))
                row.append(port_info["service_name"])
                row.append(port_info.get("product", ""))
                row.append(port_info.get("version", ""))
                ws.append(row)

        self.set_style(ws)

    def build_ip_xl(self):
        ws = self.wb.create_sheet(title="IP")
        ws.column_dimensions['A'].width = 22.0
        ws.column_dimensions['B'].width = 50.0
        ws.column_dimensions['C'].width = 10.0
        ws.column_dimensions['D'].width = 25.0
        ws.column_dimensions['E'].width = 55.0
        if self.is_ip_task:
            ws.column_dimensions['F'].width = 55.0
            column_tilte = ["IP", "端口信息", "开放端口数目", "geo", "as 编号", "操作系统"]
            ws.append(column_tilte)
            for item in get_ip_data(self.task_id):
                row = []
                row.append(item["ip"])

                port_ids = [str(x["port_id"]) for x in item["port_info"]]
                row.append(" \r\n".join(port_ids))
                row.append(len(item["port_info"]))
                if "country_name" in item["geo_city"]:
                    row.append("{}/{}".format(item["geo_city"]["country_name"],
                                              item["geo_city"]["region_name"]))
                    row.append(item["geo_asn"].get("organization", ""))
                else:
                    row.append("")
                    row.append("")

                osname = ""
                if item.get("os_info"):
                    osname = item["os_info"]["name"]
                row.append(osname)
                ws.append(row)
        else:
            ws.column_dimensions['F'].width = 60.0
            ws.column_dimensions['G'].width = 40.0
            ws.column_dimensions['H'].width = 40.0
            ws.column_dimensions['I'].width = 20.0
            column_tilte = ["IP", "端口信息", "开放端口数目", "geo", "as 编号"]
            column_tilte.append("domain")
            column_tilte.append("操作系统")
            column_tilte.append("CDN")
            column_tilte.append("类别")
            ws.append(column_tilte)
            for item in get_ip_data(self.task_id):
                row = []
                row.append(item["ip"])

                port_ids = [str(x["port_id"]) for x in item["port_info"]]
                row.append(" \r\n".join(port_ids))

                row.append(len(item["port_info"]))
                if "country_name" in item["geo_city"]:
                    row.append("{}/{}".format(item["geo_city"]["country_name"],
                                              item["geo_city"]["region_name"]))
                    row.append(item["geo_asn"].get("organization", ""))
                else:
                    row.append("")
                    row.append("")

                row.append(" \r\n".join(item.get("domain", [])))

                osname = ""
                if item.get("os_info"):
                    osname = item["os_info"]["name"]
                row.append(osname)
                row.append(item.get("cdn_name", ""))
                row.append(item.get("ip_type", ""))
                ws.append(row)

        self.set_style(ws)

    def ignore_illegal(self, content):
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        content = ILLEGAL_CHARACTERS_RE.sub(r'', content)
        return content

    def build_site_xl(self):
        ws = self.wb.active
        ws.column_dimensions['A'].width = 35.0
        ws.column_dimensions['B'].width = 40.0
        ws.column_dimensions['C'].width = 60.0
        ws.column_dimensions['D'].width = 20.0
        ws.column_dimensions['E'].width = 30.0
        ws.title = "站点"
        column_tilte = ["site", "title", "指纹", "状态码", "favicon hash"]
        ws.append(column_tilte)
        for item in get_site_data(self.task_id):
            row = []
            row.append(self.ignore_illegal(item["site"]))
            row.append(self.ignore_illegal(item["title"]))
            row.append(" \r\n".join([self.ignore_illegal(x["name"]) for x in item["finger"]]))
            row.append(item["status"])
            row.append(item["favicon"].get("hash", ""))
            ws.append(row)

        self.set_style(ws)

    def build_domain_xl(self):
        ws = self.wb.create_sheet(title="域名")
        ws.column_dimensions['A'].width = 30.0
        ws.column_dimensions['B'].width = 20.0
        ws.column_dimensions['C'].width = 50.0
        ws.column_dimensions['D'].width = 50.0

        column_tilte = ["域名", "解析类型", "记录值","关联ip"]

        ws.append(column_tilte)
        for item in get_domain_data(self.task_id):
            row = []
            row.append(item["domain"])
            row.append(item["type"])
            row.append(" \r\n".join(item["record"]))
            row.append(" \r\n".join(item["ips"]))
            ws.append(row)

        self.set_style(ws)

    def build_statist(self):
        statist = port_service_product_statist(self.task_id)
        ws = self.wb.create_sheet(title="资产统计")
        ws.column_dimensions['A'].width = 20.0
        ws.column_dimensions['F'].width = 20.0
        ws.column_dimensions['K'].width = 40.0
        ws["A1"] = "端口信息统计"
        ws["F1"] = "系统服务信息统计"
        ws["K1"] = "软件产品信息统计"

        ports = ["端口", "数量", "占比"]
        port_percent_list = statist["port_percent_list"]
        port_total = statist["port_total"]
        for port_info in port_percent_list:
            ports.append(port_info["port_id"])
            ports.append(port_info["amount"])
            ports.append(port_info["percent"])

        cnt = 0
        for row in range(5, 27):
            for col in range(1, 4):
                if cnt >= len(ports):
                    continue
                ws.cell(column=col, row=row, value=ports[cnt])
                cnt += 1

        ws["A27"] = "端口开放总数"
        ws["A28"] = port_total

        services = ["系统服务", "数量", "占比"]
        service_percent_list = statist["service_percent_list"]
        if len(service_percent_list) >= 0:
            service_total = statist["service_total"]
            for port_info in service_percent_list:
                services.append(port_info["service_name"])
                services.append(port_info["amount"])
                services.append(port_info["percent"])
            cnt = 0
            for row in range(5, 27):
                for col in range(6, 9):
                    if cnt >= len(services):
                        continue
                    ws.cell(column=col, row=row, value=services[cnt])
                    cnt += 1
            ws["F27"] = "系统服务类别总数"
            ws["F28"] = service_total

        product = ["产品", "数量", "占比"]
        product_percent_list = statist["product_percent_list"]
        if len(product_percent_list) >= 0:
            product_total = statist["product_total"]
            for port_info in product_percent_list:
                product.append(port_info["product"])
                product.append(port_info["amount"])
                product.append(port_info["percent"])
            cnt = 0
            for row in range(5, 27):
                for col in range(11, 14):
                    if cnt >= len(product):
                        continue
                    ws.cell(column=col, row=row, value=product[cnt])
                    cnt += 1
            ws["K27"] = "产品类别总数"
            ws["K28"] = product_total

        self.set_style(ws)

    def run(self):
        task_data = get_task_data(self.task_id)
        if not task_data:
            print("not found {}".format(self.task_id))
            return

        domain = task_data["target"].replace("/", "_")[:20]

        if re.findall(r"\b\d+\.\d+\.\d+\.\d+", domain):
            self.is_ip_task = True
        else:
            if task_data.get("type", "") == "ip":
                self.is_ip_task = True

        self.build_site_xl()
        self.build_ip_xl()
        self.build_service_xl()
        if not self.is_ip_task:
            self.build_domain_xl()

        self.build_statist()

        return save_virtual_workbook(self.wb)


def export_arl(task_id):
    task_id = task_id.strip()
    save = SaveTask(task_id)
    return save.run()
